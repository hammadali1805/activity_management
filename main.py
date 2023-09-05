# importing required modules 
from tkinter import *
from tkinter import ttk
from tkinter import messagebox as msg
from win32api import GetSystemMetrics
import time
import datetime
from openpyxl import load_workbook, Workbook

        


# main class having all the methods 
class App(Tk):# inherting it with T class of tkinter
    def __init__(self):
        """ constructor of the main class """
        
        super().__init__() # running constructor of parent class i.e. Tk

        try:
            self.wb=load_workbook("activities.xlsx")
            self.allActivitiesSheet=self.wb['all']
            self.declaredActivitiesSheet=self.wb['declared']
        except:
            self.wb=Workbook()
            self.allActivitiesSheet=self.wb.create_sheet('all')
            self.declaredActivitiesSheet=self.wb.create_sheet('declared')

        self.noOfActivites=self.allActivitiesSheet.max_row
        if self.allActivitiesSheet['A1'].value==None : 
            self.noOfActivites-=1
    
        self.allActivitiesName=[]
        for i in range(self.noOfActivites):
            self.allActivitiesName.append(self.allActivitiesSheet.cell(row=i+1, column=1).value)

        self.noOfDeclaredActivites=self.declaredActivitiesSheet.max_row
        if self.declaredActivitiesSheet['A1'].value==None : 
            self.noOfDeclaredActivites-=1

        self.declaredActivitiesName=[]
        for i in range(self.noOfDeclaredActivites):
            self.declaredActivitiesName.append(self.declaredActivitiesSheet.cell(row=i+1, column=1).value)

        self.leftActivitiesName=[x for x in self.allActivitiesName if x not in self.declaredActivitiesName]


        self.logo=PhotoImage(file="icon.png")
        self.activityName=StringVar()
        self.activityType=StringVar()
        self.interhouse=StringVar()
        self.activityIncharge=StringVar()
        self.ascTeacher1=StringVar()
        self.ascTeacher2=StringVar()
        self.ascTeacher3=StringVar()
        self.startDate=StringVar()
        self.startDate.set("dd/mm/yyyy")
        self.endDate=StringVar()
        self.endDate.set("dd/mm/yyyy")

        with open("teachers.txt", "r") as f:
            self.teachers=list(map(lambda x: x.strip("\n"), f.readlines()))

        self.houses=["SHIVAJI-BLUE", "ASHOKA-YELLOW", "RAMAN-RED", "TAGORE-GREEN"]
        self.resultAct=StringVar()
        self.leftAct=self.leftActivitiesName
        self.winner=StringVar()
        self.winnerHouse=StringVar()
        self.winnerClass=StringVar()
        self.firstRunner=StringVar()
        self.firstRunnerHouse=StringVar()
        self.firstRunnerClass=StringVar()
        self.secondRunner=StringVar()
        self.secondRunnerHouse=StringVar()
        self.secondRunnerClass=StringVar()
        self.actToFetch=StringVar()
        self.declaredAct=self.declaredActivitiesName


        # # displaying logo of kvs at the center of the screen before starting
        # self.geometry(f"200x150+{ GetSystemMetrics(0)//2 - 50 }+{GetSystemMetrics(1)//2 - 50 }")
        # self.overrideredirect(True)
        # self.startLabel=Label(self, image=self.logo, height=150, width=200, border=0)
        # self.startLabel.place(x=0, y=0)
        # self.update()
        # time.sleep(1)
        # self.startLabel.place_forget()
        self.layout()


    def update(self):
        super().update()

        try:
            self.wb=load_workbook("activities.xlsx")
            self.allActivitiesSheet=self.wb['all']
            self.declaredActivitiesSheet=self.wb['declared']
        except:
            self.wb=Workbook()
            self.allActivitiesSheet=self.wb.create_sheet('all')
            self.declaredActivitiesSheet=self.wb.create_sheet('declared')

        self.noOfActivites=self.allActivitiesSheet.max_row
        if self.allActivitiesSheet['A1'].value==None : 
            self.noOfActivites-=1
    
        self.allActivitiesName=[]
        for i in range(self.noOfActivites):
            self.allActivitiesName.append(self.allActivitiesSheet.cell(row=i+1, column=1).value)

        self.noOfDeclaredActivites=self.declaredActivitiesSheet.max_row
        if self.declaredActivitiesSheet['A1'].value==None : 
            self.noOfDeclaredActivites-=1

        self.declaredActivitiesName=[]
        for i in range(self.noOfDeclaredActivites):
            self.declaredActivitiesName.append(self.declaredActivitiesSheet.cell(row=i+1, column=1).value)

        self.leftActivitiesName=[x for x in self.allActivitiesName if x not in self.declaredActivitiesName]

        with open("teachers.txt", "r") as f:
            self.teachers=list(map(lambda x: x.strip("\n"), f.readlines()))

        self.leftAct=self.leftActivitiesName
        self.declaredAct=self.declaredActivitiesName


    def layout(self):
        """ method to make the layout of the window """

        self.geometry(f"{GetSystemMetrics(0)-int(0.01171875*GetSystemMetrics(0))}x{GetSystemMetrics(1)-int(0.075*GetSystemMetrics(1))}+0+0")
        self.resizable(False, False)
        self.title("  KVS Activity Manager")
        self.iconbitmap("icon.ico")
        self.overrideredirect(False)
        self.update()

        # creating menubar
        self.menubar=Menu(self)
        self.menubar.add_command(label="  HOME  ", command=self.homePage)
        self.menubar.add_command(label="  REGISTER  ", command=self.registerAct)
        self.menubar.add_command(label="  DECLARE  ", command=self.updateAct)
        self.menubar.add_command(label="  FETCH  ", command=self.fetchAct)
        self.config(menu=self.menubar)

        self.mainFrame=Frame(bg="#701e20")
        self.mainFrame.pack(expand=True, fill=BOTH)
        self.homePage()


    def clearWindow(self, bgcolor="#701e20"):
        """ this function clears the root window """

        self.mainFrame.pack_forget()
        self.mainFrame=Frame(bg=bgcolor)
        self.mainFrame.pack(expand=True, fill=BOTH)


    def homePage(self):
        self.clearWindow()

        Label(self.mainFrame, image=self.logo, height=150, width=200, border=0).place(relx=0.5, rely=0.4, anchor="center")
        Label(self.mainFrame, text="KVS ACTIVITY MANAGEMENT SOFTWARE", font="times 33 bold", border=0, bg="#701e20", fg="white").place(relx=0.5, rely=0.53, anchor="center")
        Label(self.mainFrame, text="A Software To Keep A Record Of All Activities", font="times 20 bold", border=0, bg="#701e20", fg="grey").place(relx=0.5, rely=0.6, anchor="center")



    def registerAct(self):
        self.clearWindow("#ffe2fa")


        Label(self.mainFrame, text="REGISTER A NEW ACTIVITY", fg="red", bg="#ffe2fa", font="times 35 bold underline", border=0).place(relx=0.5, rely=0.05, anchor="center")


        Label(self.mainFrame, text="ACTIVITY NAME :", bg="#ffe2fa", font="times 30 bold", border=0).place(relx=0.05, rely=0.15, anchor="w")
        Entry(self.mainFrame, textvariable=self.activityName, font="times 30 bold", fg="blue", border=0, width=35).place(relx=0.35, rely=0.15, anchor="w")


        Label(self.mainFrame, text="ACTIVITY TYPE:", bg="#ffe2fa", font="times 30 bold", border=0).place(relx=0.05, rely=0.25, anchor="w")
        Radiobutton(self.mainFrame, text="Individual", variable=self.activityType,  font="times 25 bold", bg="#ffe2fa",fg="blue", border=0, value="Individual", tristatevalue="x").place(relx=0.35, rely=0.25, anchor="w")
        Radiobutton(self.mainFrame, text="Group", variable=self.activityType,  font="times 25 bold", bg="#ffe2fa", fg="blue",  border=0, value="Group", tristatevalue="x").place(relx=0.55, rely=0.25, anchor="w")
        Radiobutton(self.mainFrame, text="Both", variable=self.activityType,  font="times 25 bold", bg="#ffe2fa", fg="blue",  border=0, value="Both", tristatevalue="x").place(relx=0.7, rely=0.25, anchor="w")


        Label(self.mainFrame, text="INTERHOUSE :", bg="#ffe2fa", font="times 30 bold", border=0).place(relx=0.05, rely=0.35, anchor="w")
        Radiobutton(self.mainFrame, text="Yes", variable=self.interhouse,  font="times 25 bold", bg="#ffe2fa", fg="blue",  border=0, value="Yes", tristatevalue="z").place(relx=0.32, rely=0.35, anchor="w")
        Radiobutton(self.mainFrame, text="No", variable=self.interhouse,  font="times 25 bold", bg="#ffe2fa", fg="blue",  border=0, value="No", tristatevalue="z").place(relx=0.45, rely=0.35, anchor="w")


        Label(self.mainFrame, text="ACTIVITY INCHARGE :", bg="#ffe2fa", font="times 30 bold", border=0).place(relx=0.05, rely=0.45, anchor="w")
        ttk.Combobox(self.mainFrame, textvariable=self.activityIncharge, width=29, values=self.teachers, font="times 30 bold", state="readonly").place(relx=0.43, rely=0.45, anchor="w")
        

        Label(self.mainFrame, text="ASSOCIATED TEACHER 1 :", bg="#ffe2fa", font="times 30 bold", border=0).place(relx=0.05, rely=0.55, anchor="w")
        ttk.Combobox(self.mainFrame, textvariable=self.ascTeacher1, width=26, values=self.teachers, font="times 30 bold", state="readonly").place(relx=0.48, rely=0.55, anchor="w")


        Label(self.mainFrame, text="ASSOCIATED TEACHER 2 :", bg="#ffe2fa", font="times 30 bold", border=0).place(relx=0.05, rely=0.65, anchor="w")
        ttk.Combobox(self.mainFrame, textvariable=self.ascTeacher2, width=26, values=self.teachers, font="times 30 bold", state="readonly").place(relx=0.48, rely=0.65, anchor="w")


        Label(self.mainFrame, text="ASSOCIATED TEACHER 3 :", bg="#ffe2fa", font="times 30 bold", border=0).place(relx=0.05, rely=0.75, anchor="w")
        ttk.Combobox(self.mainFrame, textvariable=self.ascTeacher3, width=26, values=self.teachers, font="times 30 bold", state="readonly").place(relx=0.48, rely=0.75, anchor="w")


        Label(self.mainFrame, text="STARTING DATE:", bg="#ffe2fa", font="times 30 bold", border=0).place(relx=0.05, rely=0.85, anchor="w")
        Entry(self.mainFrame, textvariable=self.startDate, font="times 30 bold", fg="blue", border=0, width=11).place(relx=0.34, rely=0.85, anchor="w")


        Label(self.mainFrame, text="ENDING DATE:", bg="#ffe2fa", font="times 30 bold", border=0).place(relx=0.52, rely=0.85, anchor="w")
        Entry(self.mainFrame, textvariable=self.endDate, font="times 30 bold", fg="blue", border=0, width=11).place(relx=0.77, rely=0.85, anchor="w")


        Button(self.mainFrame, text="REGISTER", command=self.registerInFile, bg="light grey", font="times 25 bold").place(relx=0.5, rely=0.95, anchor="center")


    def dateValidator(self, inputDate):
        try:
            day, month, year = inputDate.split("/")
        except:
            return False

        isvalidDate = True
        try:
            datetime.datetime(int(year), int(month), int(day))
        except:
            isvalidDate = False

        return isvalidDate

    
    def compareDate(self, sD, eD):
        date={}
        x=0
        for i in [sD, eD]:
            x+=1
            day, month, year = i.split("/")
            date[x]=datetime.datetime(int(year), int(month), int(day))

        if date[1]>date[2]:
            return True
        else:
            return False


    def registerInFile(self):

        data=[
        self.activityName.get(),
        self.activityType.get(),
        self.interhouse.get(),
        self.activityIncharge.get(),
        self.ascTeacher1.get(),
        self.ascTeacher2.get(),
        self.ascTeacher3.get(),
        self.startDate.get(),
        self.endDate.get()
        ]

        for i in data:
            if i.strip()=="":
                msg.showerror("Error!", "One or more fields are Empty!")
                return

        if not self.dateValidator(data[-1]) or not self.dateValidator(data[-2]):
            msg.showerror("Error!", "Please Enter a valid date!") 
            return

        if len(data[-1])!=10 or len(data[-2])!=10:
            msg.showerror("Error!", "Please check the format of date!") 
            return
        
        if self.compareDate(data[-2], data[-1]):
            msg.showerror("Error!", "Starting Date can't be greater than Ending Date!") 
            return
            
        rowToPlace=self.noOfActivites+1
        for i in range(len(data)):
            self.allActivitiesSheet.cell(column=i+1, row=rowToPlace).value=data[i]

        self.wb.save("activities.xlsx")
        msg.showinfo("Success!", "Activity has been registered successfully!")

        self.activityName.set("")
        self.activityType.set("")
        self.interhouse.set("")
        self.activityIncharge.set("")
        self.ascTeacher1.set("")
        self.ascTeacher2.set("")
        self.ascTeacher3.set("")
        self.startDate.set("dd/mm/yyyy")
        self.endDate.set("dd/mm/yyyy")

        self.update()
        
    def updateAct(self):
        self.clearWindow("#f6ffa8")

        Label(self.mainFrame, text="DECLARE RESULT OF AN ACTIVITY", fg="red", bg="#f6ffa8", font="times 35 bold underline", border=0).place(relx=0.5, rely=0.05, anchor="center")


        Label(self.mainFrame, text="ACTIVITY NAME :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.05, rely=0.15, anchor="w")
        ttk.Combobox(self.mainFrame, textvariable=self.resultAct, width=35, values=self.leftAct, font="times 30 bold", state="readonly").place(relx=0.35, rely=0.15, anchor="w")


        Label(self.mainFrame, text="Winner :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.05, rely=0.22, anchor="w")
        Label(self.mainFrame, text="Name :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.25, rely=0.27, anchor="w")
        Entry(self.mainFrame, textvariable=self.winner, font="times 30 bold", border=0, fg="blue", width=34).place(relx=0.38, rely=0.27, anchor="w")
        Label(self.mainFrame, text="House :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.25, rely=0.37, anchor="w")
        ttk.Combobox(self.mainFrame, textvariable=self.winnerHouse, width=20, values=self.houses, font="times 30 bold", state="readonly").place(relx=0.38, rely=0.37, anchor="w")
        Label(self.mainFrame, text="Class :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.75, rely=0.37, anchor="w")
        ttk.Combobox(self.mainFrame, textvariable=self.winnerClass, width=3, values=["1", '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'], font="times 30 bold", state="readonly").place(relx=0.85, rely=0.37, anchor="w")


        Label(self.mainFrame, text="First Runner :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.05, rely=0.45, anchor="w")
        Label(self.mainFrame, text="Name :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.25, rely=0.5, anchor="w")
        Entry(self.mainFrame, textvariable=self.firstRunner, font="times 30 bold", border=0, fg="blue", width=34).place(relx=0.38, rely=0.5, anchor="w")
        Label(self.mainFrame, text="House :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.25, rely=0.6, anchor="w")
        ttk.Combobox(self.mainFrame, textvariable=self.firstRunnerHouse, width=20, values=self.houses, font="times 30 bold", state="readonly").place(relx=0.38, rely=0.6, anchor="w")
        Label(self.mainFrame, text="Class :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.75, rely=0.6, anchor="w")
        ttk.Combobox(self.mainFrame, textvariable=self.firstRunnerClass, width=3, values=["1", '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'], font="times 30 bold", state="readonly").place(relx=0.85, rely=0.6, anchor="w")
        

        Label(self.mainFrame, text="Second Runner :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.05, rely=0.69, anchor="w")
        Label(self.mainFrame, text="Name :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.25, rely=0.74, anchor="w")
        Entry(self.mainFrame, textvariable=self.secondRunner, font="times 30 bold", border=0, fg="blue", width=34).place(relx=0.38, rely=0.74, anchor="w")
        Label(self.mainFrame, text="House :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.25, rely=0.84, anchor="w")
        ttk.Combobox(self.mainFrame, textvariable=self.secondRunnerHouse, width=20, values=self.houses, font="times 30 bold", state="readonly").place(relx=0.38, rely=0.84, anchor="w")
        Label(self.mainFrame, text="Class :", bg="#f6ffa8", font="times 30 bold", border=0).place(relx=0.75, rely=0.84, anchor="w")
        ttk.Combobox(self.mainFrame, textvariable=self.secondRunnerClass, width=3, values=["1", '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'], font="times 30 bold", state="readonly").place(relx=0.85, rely=0.84, anchor="w")


        Button(self.mainFrame, text="DECLARE", command=self.updateInFile, bg="light grey", font="times 25 bold").place(relx=0.5, rely=0.95, anchor="center")


    def updateInFile(self):

        data=[
        self.resultAct.get(),
        self.winner.get(),
        self.winnerHouse.get(),
        self.winnerClass.get(),
        self.firstRunner.get(),
        self.firstRunnerHouse.get(),
        self.firstRunnerClass.get(),
        self.secondRunner.get(),
        self.secondRunnerHouse.get(),
        self.secondRunnerClass.get()
        ]

        for i in data:
            if i.strip()=="":
                msg.showerror("Error!", "One or more fields are Empty!")
                return

        rowToPlace=self.noOfDeclaredActivites+1
        for i in range(len(data)):
            self.declaredActivitiesSheet.cell(column=i+1, row=rowToPlace).value=data[i]

        self.wb.save("activities.xlsx")
        msg.showinfo("Success!", "Result of the activity has been declared successfully!")


        self.resultAct.set("")
        self.winner.set("")
        self.winnerHouse.set("")
        self.winnerClass.set("")
        self.firstRunner.set("")
        self.firstRunnerHouse.set("")
        self.firstRunnerClass.set("")
        self.secondRunner.set("")
        self.secondRunnerHouse.set("")
        self.secondRunnerClass.set("")

        self.update()

    def fetchAct(self):
        self.clearWindow("light blue")

        
        Label(self.mainFrame, text="FETCH ENTRY OF AN ACTIVITY", fg="red", bg="light blue", font="times 35 bold underline", border=0).place(relx=0.5, rely=0.05, anchor="center")


        Label(self.mainFrame, text="ACTIVITY NAME :", bg="light blue", font="times 35 bold", border=0).place(relx=0.5, rely=0.2, anchor="center")
        ttk.Combobox(self.mainFrame, textvariable=self.actToFetch, width=40, values=self.declaredAct, font="times 30 bold", state="readonly").place(relx=0.5, rely=0.4, anchor="center")

       
        Label(self.mainFrame, text="PLEASE SELECT THE ACTIVITY FIRST\n THEN CLICK ON FETCH BUTTON\n ", bg="light blue", font="times 30 bold", border=0).place(relx=0.5, rely=0.65, anchor="center")


        Button(self.mainFrame, text="FETCH", command=self.fetchFromFile, bg="light grey", font="times 25 bold").place(relx=0.5, rely=0.85, anchor="center")

    
    def fetchFromFile(self):

        if self.actToFetch.get().strip()=="":
            msg.showerror("Error!", "Please Select Any Activity First!")
            return

        self.clearWindow("light blue")

        a=StringVar()
        b=StringVar()
        c=StringVar()
        d=StringVar()
        e=StringVar()
        f=StringVar()
        g=StringVar()
        h=StringVar()
        i=StringVar()
        j=StringVar()
        k=StringVar()
        l=StringVar()
        m=StringVar()
        n=StringVar()
        o=StringVar()
        p=StringVar()
        q=StringVar()
        r=StringVar()

        Label(self.mainFrame, text="FETCHED ENTRY OF THE ACTIVITY", fg="red", bg="light blue", font="times 35 bold underline", border=0).place(relx=0.5, rely=0.05, anchor="center")


        Label(self.mainFrame, text="ACTIVITY NAME :", bg="light blue", font="times 25 bold", border=0).place(relx=0.05, rely=0.15, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", border=0, width=45, textvariable=a, fg="blue").place(relx=0.3, rely=0.15, anchor="w")


        Label(self.mainFrame, text="ACTIVITY TYPE :", bg="light blue", font="times 25 bold", border=0).place(relx=0.05, rely=0.22, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=25, textvariable=b).place(relx=0.3, rely=0.22, anchor="w")


        Label(self.mainFrame, text="INTERHOUSE :", bg="light blue", font="times 25 bold", border=0).place(relx=0.65, rely=0.22, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=4, textvariable=c).place(relx=0.85, rely=0.22, anchor="w")


        Label(self.mainFrame, text="ACTIVITY INCHARGE :", bg="light blue", font="times 25 bold", border=0).place(relx=0.05, rely=0.29, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=41, textvariable=d).place(relx=0.355, rely=0.29, anchor="w")


        Label(self.mainFrame, text="ASOCIATED TEACHER 1 :", bg="light blue", font="times 25 bold", border=0).place(relx=0.05, rely=0.36, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=39, textvariable=e).place(relx=0.384, rely=0.36, anchor="w")


        Label(self.mainFrame, text="ASOCIATED TEACHER 2 :", bg="light blue", font="times 25 bold", border=0).place(relx=0.05, rely=0.43, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=39, textvariable=f).place(relx=0.384, rely=0.43, anchor="w")


        Label(self.mainFrame, text="ASOCIATED TEACHER 3 :", bg="light blue", font="times 25 bold", border=0).place(relx=0.05, rely=0.5, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=39, textvariable=g).place(relx=0.384, rely=0.5, anchor="w")


        Label(self.mainFrame, text="STARTED ON:", bg="light blue", font="times 25 bold", border=0).place(relx=0.05, rely=0.57, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=11, textvariable=h).place(relx=0.28, rely=0.57, anchor="w")


        Label(self.mainFrame, text="ENDED ON:", bg="light blue", font="times 25 bold", border=0).place(relx=0.56, rely=0.57, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=11, textvariable=i).place(relx=0.76, rely=0.57, anchor="w")


        Label(self.mainFrame, text="NAME", bg="light blue", font="times 25 bold underline", border=0).place(relx=0.3, rely=0.64, anchor="w")
        Label(self.mainFrame, text="HOUSE", bg="light blue", font="times 25 bold underline", border=0).place(relx=0.58, rely=0.64, anchor="w")
        Label(self.mainFrame, text="CLASS", bg="light blue", font="times 25 bold underline", border=0).place(relx=0.8, rely=0.64, anchor="w")


        Label(self.mainFrame, text="WINNER :", bg="light blue", font="times 25 bold", border=0).place(relx=0.05, rely=0.71, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=20, textvariable=j).place(relx=0.2, rely=0.71, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=20, textvariable=k).place(relx=0.5, rely=0.71, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=3, textvariable=l).place(relx=0.82, rely=0.71, anchor="w")


        Label(self.mainFrame, text="I RUNNER :", bg="light blue", font="times 25 bold", border=0).place(relx=0.05, rely=0.78, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=20, textvariable=m).place(relx=0.2, rely=0.78, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=20, textvariable=n
        ).place(relx=0.5, rely=0.78, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=3, textvariable=o).place(relx=0.82, rely=0.78, anchor="w")



        Label(self.mainFrame, text="II RUNNER :", bg="light blue", font="times 25 bold", border=0).place(relx=0.04, rely=0.85, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=20, textvariable=p).place(relx=0.2, rely=0.85, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=20, textvariable=q).place(relx=0.5, rely=0.85, anchor="w")
        Entry(self.mainFrame, font="times 25 bold", fg="blue", border=0, width=3, textvariable=r).place(relx=0.82, rely=0.85, anchor="w")


        Button(self.mainFrame, text="FETCH ANOTHER", command=self.fetchAct, bg="light grey", font="times 25 bold").place(relx=0.3, rely=0.95, anchor="center")

        Button(self.mainFrame, text="DELETE THIS RECORD", command=self.dltAct, bg="light grey", font="times 25 bold").place(relx=0.7, rely=0.95, anchor="center")

        row_sheet1=self.allActivitiesName.index(self.actToFetch.get())+1
        x=0
        for i in [a, b, c, d, e, f, g, h, i]:
            x+=1
            i.set(self.allActivitiesSheet.cell(column=x, row=row_sheet1).value)

        row_sheet2=self.declaredActivitiesName.index(self.actToFetch.get())+1
        x=1
        for i in [j, k, l, m, n, o, p, q, r]:
            x+=1
            i.set(self.declaredActivitiesSheet.cell(column=x, row=row_sheet2).value)


    def dltAct(self):

        row_sheet1=self.allActivitiesName.index(self.actToFetch.get())+1
        self.allActivitiesSheet.delete_rows(row_sheet1)

        row_sheet2=self.declaredActivitiesName.index(self.actToFetch.get())+1
        self.declaredActivitiesSheet.delete_rows(row_sheet2)

        self.wb.save("activities.xlsx")
        msg.showinfo("Success!", "Activity has been deleted successfully!")

        self.actToFetch=""
        self.update()
        self.fetchAct()


if __name__=="__main__":
    root=App()     
    root.mainloop()